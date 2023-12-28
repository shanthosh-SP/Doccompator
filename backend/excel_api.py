import os
import pandas as pd
import openpyxl
import re
import nltk
import torch
from inscriptis import get_text
from sklearn.metrics.pairwise import cosine_similarity
from transformers import BertTokenizer, BertModel
from tabulate import tabulate
import textwrap
#nltk.download('punkt') 

from nltk import download


# Disable SSL verification
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

# Download 'punkt'
download('punkt')


def extract_text_from_excel(input_excel_file, input_html_file):
    with open(input_html_file, 'r', encoding="ISO-8859-1") as file:
        html_content = file.read()

    # Assuming you have a function get_text() to extract text from HTML
    text_content = get_text(html_content)

    # Load the Excel file
    excel_file = input_excel_file
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook.active

    # Define a maximum width for text wrapping
    max_width = 50  # Adjust as needed

    # Initialize a list to store the extracted data
    data = []

    # Extract column headers from the first row
    headers = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        cell_text = cell.value if cell.value else ''
        wrapped_text = textwrap.fill(str(cell_text), width=max_width)
        headers.append(wrapped_text)

    # Append the headers to the data list
    data.append(headers)

    # Iterate through the rows and columns in the Excel sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_data = []
        for cell in row:
            cell_text = cell.value if cell.value else ''
            wrapped_text = textwrap.fill(str(cell_text), width=max_width)
            row_data.append(wrapped_text)
        data.append(row_data)

    # Combine the headers and data and format them with tabulate
    table = tabulate(data, tablefmt="grid", headers="firstrow")

    # Read all sheets from the Excel file into a dictionary of DataFrames
    # Create an empty DataFrame to store the data from all sheets
    all_sheets_data = pd.DataFrame()

    # Read all sheets from the Excel file into a dictionary of DataFrames
    xls = pd.ExcelFile(input_excel_file)
    sheets_data = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}

    def save_text_to_file(sheets_data, output_file):
        # Create or open the text file for writing
        with open(output_file, 'w') as text_file:
            # Iterate through all sheets and write data to the text file
            for sheet_name, sheet_data in sheets_data.items():
                text_file.write(f"{sheet_name}\n")
                sheet_data = sheet_data.fillna('')
                text_file.write(sheet_data.to_string(index=False, header=True))
                text_file.write("\n\n")

    input_excel_file = os.path.splitext(input_excel_file)[0].strip()
    output_excel_file = f"{input_excel_file}_extracted.txt"

    if sheets_data:
        save_text_to_file(sheets_data, output_excel_file)
    else:
        print("Error: The number of extracted pages does not match the Excel metadata.")

    excel_text = ""
    with open(output_excel_file, 'r', encoding="ISO-8859-1") as excel_file:
        excel_text = excel_file.read().lower()

    return {'excel_extraction_tabulated': table, 'excel_extraction_plain': excel_text, 'html': text_content}

def compare_excel_with_html(excel_text, html_text):
    try:
        excel_words = set(nltk.word_tokenize(excel_text))
        html_words = set(nltk.word_tokenize(html_text))

        # Finding the difference
        difference_words = excel_words - html_words
        for word in difference_words:
            excel_text = re.sub(rf"(?<!>)\b({re.escape(word)})\b(?!<)", r"<span style='background-color: red;'>\1</span>", ppt_text, flags=re.IGNORECASE)

        # Finding the line, position, page
        word_positions = {}
        excel_lines= excel_text.splitlines()

        page_number = 0
        line_number = 0

        for line in excel_lines:
            if line.startswith("sheet"):
                match = re.match(r'sheet(\d+)', line)
                if match:
                    page_number = int(match.group(1))
                line_number = 0
            else:
                line_number += 1

            line_words = list(re.findall(r'\b\w+\b', line))
            for position, word in enumerate(line_words):
                
                if line_words.startswith("sheet"):
                    continue 
                if len(word) > 1 and word.lower() in difference_words:

                    if word in difference_words:
                        if word not in word_positions:
                            word_positions[word] = []
                        word_positions[word].append({
                            'Page': page_number,
                            'Line': line_number,
                            'Position': position,
                            'LineContent': line
                        })
        for word, positions in word_positions.items():
            unique_positions = []
            seen_positions = set()
        for position in positions:
            position_tuple = (position['Page'], position['Line'], position['Position'])
            if position_tuple not in seen_positions:
                seen_positions.add(position_tuple)
                unique_positions.append(position)
        word_positions[word] = unique_positions
        
        output = f"excel_htmlcompare1.txt"
        with open(output, 'w', encoding='utf-8') as result_file:
            for word, positions in word_positions.items():
                for data in positions:
                    result_file.write(f"Word: {word}\n")
                    result_file.write(f"Page: {data['Page']}\n")
                    result_file.write(f"Line: {data['Line']}\n")
                    result_file.write(f"Position: {data['Position']}\n")
                    result_file.write(f"Line Content: {data['LineContent']}\n\n")

        # Comparison using Jaccard Similarity
        jaccard_similarity = len(excel_words.intersection(html_words)) / len(excel_words.union(html_words))
        percentage_difference = (1 - jaccard_similarity) * 100

        # Comparison using BERT Cosine Similarity
        model_name = "bert-base-uncased"
        tokenizer = BertTokenizer.from_pretrained(model_name)
        model = BertModel.from_pretrained(model_name)

        tokens1 = tokenizer(excel_text, return_tensors='pt', padding=True, truncation=True)
        tokens2 = tokenizer(html_text, return_tensors='pt', padding=True, truncation=True)

        with torch.no_grad():
            embeddings1 = model(**tokens1).last_hidden_state.mean(dim=1)
            embeddings2 = model(**tokens2).last_hidden_state.mean(dim=1)

        similarity = cosine_similarity(embeddings1, embeddings2)
        bert_cosine_similarity = similarity[0][0].item()  # Convert float32 to a standard Python float

        # Include extracted PDF and HTML texts and comparison output in the response
        with open(output, 'r', encoding='utf-8') as result_file:
            output_content = result_file.read()

        response_data = {
            "bert_cosine_similarity": float(similarity[0][0]),
            "pdf_text": excel_text,
            "html_text": html_text,
            "comparison_output": {
                "file_path": output,
                "content": output_content
            }
        }

        return response_data
    
    except Exception as e:
        return {"error": str(e)}


def process_files_and_return_results(excel_file_path, html_file_path):
    # Extract text from Excel
    excel_extraction_results = extract_text_from_excel(excel_file_path, html_file_path)

    # Compare Excel with HTML
    compare_results = compare_excel_with_html(excel_extraction_results['excel_extraction_plain'],excel_extraction_results['html'])

    return compare_results


# pdf_api.py
import logging
import subprocess
from tika import parser
import os
from inscriptis import get_text
# import nltk
# nltk.download('punkt') 
import re
import torch
from transformers import BertTokenizer, BertModel
from sklearn.metrics.pairwise import cosine_similarity
from flask import jsonify  # Import jsonify for JSON response
import nltk
from nltk import download
import os
import pandas as pd
import openpyxl
import re
import nltk
import torch
from inscriptis import get_text
from sklearn.metrics.pairwise import cosine_similarity
from transformers import BertTokenizer, BertModel
from tabulate import tabulate
import textwrap
#nltk.download('punkt') 

from nltk import download


# Disable SSL verification
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

# Download 'punkt'
download('punkt')

# Disable SSL verification
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

# Download 'punkt'
download('punkt')

def extract_text_with_layout(file_path):
    with open(file_path, 'r', encoding='utf-8') as text_file:
        text = text_file.read()
    return text

def process_pdf_and_html(input_pdf_file, input_html_file):
    # Process PDF
    output_text_file = 'output.txt'
    subprocess.run(['pdftotext', '-layout', input_pdf_file, output_text_file])

    parsed_pdf = parser.from_file(input_pdf_file)
    num_pages = int(parsed_pdf['metadata'].get('xmpTPg:NPages', 0))

    extracted_text = extract_text_with_layout(output_text_file)
    pages = extracted_text.split('\x0c')
    pages_with_numbers = []

    for num_pages, page_text in enumerate(pages, 1):
        if page_text.strip():
            pages_with_numbers.append(f"Page {num_pages}\n{page_text}")

    combined_text = '\n'.join(pages_with_numbers)

    # Process HTML
    if not os.path.exists(input_html_file):
        print(f'The file {input_html_file} does not exist.')
        exit()

    with open(input_html_file, 'r', encoding="ISO-8859-1") as file:
        html_content = file.read()

    # Assuming you have a function get_text() to extract text from HTML
    text_content = get_text(html_content)

    return combined_text, text_content


def process_pdf_and_excel(input_pdf_file, input_excel_file):
    # Process PDF
    output_text_file = 'output.txt'
    subprocess.run(['pdftotext', '-layout', input_pdf_file, output_text_file])

    parsed_pdf = parser.from_file(input_pdf_file)
    num_pages = int(parsed_pdf['metadata'].get('xmpTPg:NPages', 0))

    extracted_text = extract_text_with_layout(output_text_file)
    pages = extracted_text.split('\x0c')
    pages_with_numbers = []

    for num_pages, page_text in enumerate(pages, 1):
        if page_text.strip():
            pages_with_numbers.append(f"Page {num_pages}\n{page_text}")

    combined_text = '\n'.join(pages_with_numbers)

    # # Process HTML
    # if not os.path.exists(input_html_file):
    #     print(f'The file {input_html_file} does not exist.')
    #     exit()

    # with open(input_html_file, 'r', encoding="ISO-8859-1") as file:
    #     html_content = file.read()

    # # Assuming you have a function get_text() to extract text from HTML
    # text_content = get_text(html_content)
        # Load the Excel file
    excel_file = input_excel_file
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook.active

    # Define a maximum width for text wrapping
    max_width = 50  # Adjust as needed

    # Initialize a list to store the extracted data
    data = []

    # Extract column headers from the first row
    headers = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        cell_text = cell.value if cell.value else ''
        wrapped_text = textwrap.fill(str(cell_text), width=max_width)
        headers.append(wrapped_text)

    # Append the headers to the data list
    data.append(headers)

    # Iterate through the rows and columns in the Excel sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_data = []
        for cell in row:
            cell_text = cell.value if cell.value else ''
            wrapped_text = textwrap.fill(str(cell_text), width=max_width)
            row_data.append(wrapped_text)
        data.append(row_data)

    # Combine the headers and data and format them with tabulate
    table = tabulate(data, tablefmt="grid", headers="firstrow")

    # Read all sheets from the Excel file into a dictionary of DataFrames
    # Create an empty DataFrame to store the data from all sheets
    all_sheets_data = pd.DataFrame()

    # Read all sheets from the Excel file into a dictionary of DataFrames
    xls = pd.ExcelFile(input_excel_file)
    sheets_data = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}

    def save_text_to_file(sheets_data, output_file):
        # Create or open the text file for writing
        with open(output_file, 'w') as text_file:
            # Iterate through all sheets and write data to the text file
            for sheet_name, sheet_data in sheets_data.items():
                text_file.write(f"{sheet_name}\n")
                sheet_data = sheet_data.fillna('')
                text_file.write(sheet_data.to_string(index=False, header=True))
                text_file.write("\n\n")

    input_excel_file = os.path.splitext(input_excel_file)[0].strip()
    output_excel_file = f"{input_excel_file}_extracted.txt"

    if sheets_data:
        save_text_to_file(sheets_data, output_excel_file)
    else:
        print("Error: The number of extracted pages does not match the Excel metadata.")

    excel_text = ""
    with open(output_excel_file, 'r', encoding="ISO-8859-1") as excel_file:
        excel_text = excel_file.read()
    print("22222222222222222222222",combined_text)
    print("22222222222222222222222",excel_text)

    return combined_text,  excel_text


def compare_pdf_with_html(pdf_text, html_text):
    try:
        # Tokenize the PDF and HTML texts
        from nltk.tokenize import regexp_tokenize 
        pdf_words = set(regexp_tokenize(pdf_text, "[\w']+"))
        html_words = set(regexp_tokenize(html_text, "[\w']+"))

        # Finding the difference
        difference_words = pdf_words - html_words
        difference_words = {word for word in difference_words if word.lower() != 'page' and not word.startswith('Page') and len(word) > 3}

        pdf_text_span = pdf_text
        for word in difference_words:
            pdf_text_span = re.sub(rf"(?<!>)\b({re.escape(word)})\b(?!<)", r'\1', pdf_text_span, flags=re.IGNORECASE)

        # Finding the line, position, and page
        word_positions = {}
        seen_positions = set()
        # Initialize a counter for generating unique serial numbers

        pdf_lines = pdf_text.splitlines()
        page_number = 0
        line_number = 0

        for line in pdf_lines:
            if line.startswith("Page"):
                match = re.match(r'Page (\d+)', line)
                if match:
                    page_number = int(match.group(1))
                line_number = 0
            else:
                line_number += 1

            line_words = list(regexp_tokenize(line, "[\w,']+"))
            
            for position, word in enumerate(line_words):
                if len(word) <= 2:
                    continue
                
                if word in difference_words:
                    if word not in word_positions:
                        word_positions[word] = []

                    position_tuple = (page_number, line_number, position)
                    if position_tuple not in seen_positions:
                        seen_positions.add(position_tuple)

                        word_positions[word].append({
                           
                            'Page': page_number,
                            'Line': line_number,
                            'Position': position,
                            'LineContent': line
                        })

         # Increment the serial number counter

        # Save the comparison output to a file
        output = "pdf_htmlcompare.txt"
        with open(output, 'w', encoding='utf-8') as result_file:
            for word, positions in word_positions.items():
                for data in positions:
                    result_file.write(f"Word: {word}\n")
                    result_file.write(f"Page: {data['Page']}\n")
                    result_file.write(f"Line: {data['Line']}\n")
                    result_file.write(f"Position: {data['Position']}\n")
                    result_file.write(f"Line Content: {data['LineContent']}\n\n")

        logging.info(f"Words in PDF but not in HTML, along with positions, saved to {output}")


        # Comparison using Jaccard Similarity
        jaccard_similarity = len(pdf_words.intersection(html_words)) / len(pdf_words.union(html_words))
        percentage_difference = (1 - jaccard_similarity) * 100
        logging.info(f"Jaccard Similarity Score: {jaccard_similarity:.2f}")

        # Comparison using BERT Cosine Similarity
        model_name = "bert-base-uncased"
        tokenizer = BertTokenizer.from_pretrained(model_name)
        model = BertModel.from_pretrained(model_name)

        tokens1 = tokenizer(pdf_text, return_tensors='pt', padding=True, truncation=True)
        tokens2 = tokenizer(html_text, return_tensors='pt', padding=True, truncation=True)

        with torch.no_grad():
            embeddings1 = model(**tokens1).last_hidden_state.mean(dim=1)
            embeddings2 = model(**tokens2).last_hidden_state.mean(dim=1)

        similarity = cosine_similarity(embeddings1, embeddings2)
        bert_cosine_similarity = similarity[0][0].item()  # Convert float32 to a standard Python float
        logging.info(f"Bert Cosine Similarity: {bert_cosine_similarity:.2f}")

        # Include extracted PDF and HTML texts and comparison output in the response
        with open(output, 'r', encoding='utf-8') as result_file:
            output_content = result_file.read()

        response_data = {
            "bert_cosine_similarity": float(similarity[0][0]),
            "jaccard_similarity": float(jaccard_similarity),
            "pdf_text": pdf_text_span,
            "html_text": html_text,
            "comparison_output": {
                "file_path": output,
                "content": output_content
            }
        }

        return response_data
    
    except Exception as e:
        return {"error": str(e)}



# ... (other functions if any)
